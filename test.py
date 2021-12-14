from openpyxl import load_workbook

xl = load_workbook('pl.xlsx')
edit = xl['Sheet1']
x = (edit['A1'])
y = (input('numero pa'))
r = [edit.cell(row=i, column=2).value for i in range(1, 100)]
z = 55

print(y+'g')


#x =edit['A1']
#y =edit['c1']
#r = [edit.cell(row=i,column=2).value for i in range(1,11)]
#edit.cell(row=(x.value), column=1).value = 455
#if y.value in r:
#    print("ya esta")
#else:
#    print('sad')
#print(y.value)
#print(x.value)
#print(r)

